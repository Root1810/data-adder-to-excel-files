import openpyxl

file_path = 'file.xlsx'
workbook = openpyxl.load_workbook(file_path)

sheet_names = workbook.sheetnames

print("Available sheets:")
for idx, sheet_name in enumerate(sheet_names, start=1):
    print(f"{idx}. {sheet_name}")

selected_index = int(input("\nSelect the sheet number you want to edit: ")) - 1
if selected_index < 0 or selected_index >= len(sheet_names):
    print("Invalid selection. Please run the script again and select a valid sheet number.")
else:
    selected_sheet_name = sheet_names[selected_index]
    sheet = workbook[selected_sheet_name]

    column_headers = [cell.value for cell in sheet[1]]

    print(f"\nColumns in the sheet '{selected_sheet_name}':")
    for col_idx, header in enumerate(column_headers, start=1):
        print(f"{col_idx}. {header}")

    selected_col_index = int(input("\nSelect the column number you want to add data to: ")) - 1
    if selected_col_index < 0 or selected_col_index >= len(column_headers):
        print("Invalid column selection. Please run the script again and select a valid column number.")
    else:
        selected_column_name = column_headers[selected_col_index]
        column_letter = openpyxl.utils.get_column_letter(selected_col_index + 1)
        print(f"\nYou selected the column: '{selected_column_name}' (Column: {column_letter})")

        existing_values = set()
        for row in range(2, sheet.max_row + 1):
            cell_value = sheet[f"{column_letter}{row}"].value
            if cell_value:
                existing_values.add(cell_value.strip())

        txt_file_path = 'data_add.txt'
        with open(txt_file_path, 'r') as file:
            data = file.readlines()

        added_any = False
        for line in data:
            value = line.strip()

            if value in existing_values:
                print(f"{value} already exists in the column.")
            else:
                row_idx = 2
                while sheet[f"{column_letter}{row_idx}"].value:
                    row_idx += 1
                sheet[f"{column_letter}{row_idx}"] = value
                existing_values.add(value)
                added_any = True

        if added_any:
            workbook.save(file_path)
            print(f"\nData added to the '{selected_column_name}' column (Column {column_letter}) in sheet '{selected_sheet_name}' and saved successfully.")
        else:
            print("\nNo new data was added as all entries already exist.")
